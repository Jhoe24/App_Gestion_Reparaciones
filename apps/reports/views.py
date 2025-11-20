from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, CreateView, DetailView
from apps.authentication.mixins import AdminRequiredMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required
from apps.authentication.decorators import tech_required
from django.views.decorators.http import require_http_methods
from .forms import FichaEntradaForm, SeguimientoForm
from django.contrib import messages
from .models import FichaEntrada, Seguimiento, QueuedEmail
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist
import json
from django.template.loader import render_to_string
from django.db.models import Q, Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import datetime
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment
import io
import os
import logging
from django.core.files.storage import default_storage
from uuid import uuid4
from django.db import transaction
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.core.mail import send_mail
from django.utils.html import strip_tags

logger = logging.getLogger(__name__)



CustomUser = get_user_model()


class AdminReportsDashboardView(AdminRequiredMixin, TemplateView):
    template_name = 'reports/admin_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['reportes'] = 0
        context['reportes_pendientes'] = 0
        context['reportes_resueltos'] = 0
        return context

class UserReportsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/user_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['reportes'] = 0 #Reporte.objects.filter(usuario_reporta=self.request.user).order_by('-fecha_reporte')
        return context

class ReporteCreateView(LoginRequiredMixin, CreateView):
    model = None
    form_class = None
    template_name = 'reports/reporte_form.html'
    success_url = reverse_lazy('reports:user_dashboard')

    def form_valid(self, form):
        form.instance.usuario_reporta = self.request.user
        return super().form_valid(form)

class ReporteDetailView(LoginRequiredMixin, DetailView):
    model = None
    template_name = 'reports/reporte_detail.html'
    context_object_name = 'reporte'

    def dispatch(self, request, *args, **kwargs):
        reporte = self.get_object()
        if request.user.is_superuser or getattr(request.user, 'rol', None) == 'administrador':
            return super().dispatch(request, *args, **kwargs)
        if reporte.usuario_reporta == request.user:
            return super().dispatch(request, *args, **kwargs)
        raise PermissionDenied('No tienes permiso para ver este reporte.')


@tech_required
def ficha_entrada_view(request):
    try:
        if request.method == 'POST':
            form = FichaEntradaForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Ficha de entrada creada exitosamente.")
                return redirect('reports:ficha_entrada')
        else:
            form = FichaEntradaForm()
            messages.info(request, "Por favor, completa el formulario para crear una ficha de entrada.")
        return render(request, 'reports/ficha_entrada.html', {'form': form})
    except Exception as e:
        messages.error(request, f"Ocurrió un error al procesar la ficha de entrada: {e}")
        return render(request, 'reports/ficha_entrada.html', {'form': FichaEntradaForm()})

@tech_required
def historial_equipos(request):
    search_query = request.GET.get('search', '')
    # Base queryset
    registros = FichaEntrada.objects.all()

    # Si el usuario es técnico (y no es superuser), limitar a sus fichas asignadas
    is_tech_user = (getattr(request.user, 'rol', None) == 'tecnico') and (not request.user.is_superuser)
    if is_tech_user:
        registros = registros.filter(tecnico_asignado=request.user)

    # Aplicar búsqueda sobre el queryset ya filtrado
    if search_query:
        # agregamos el filtro por el tecnico asignado
        registros = registros.filter(
            Q(codigo__icontains=search_query) |
            Q(nombre_cliente__icontains=search_query) |
            Q(apellido_cliente__icontains=search_query) |
            Q(cedula_cliente__icontains=search_query) |
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(dependencia__icontains=search_query) |
            Q(ubicacion__icontains=search_query) |
            Q(departamento_cliente__icontains=search_query) |
            Q(numero_serie__icontains=search_query) |
            Q(descripcion_falla__icontains=search_query) |
            Q(tipo_falla__icontains=search_query) |
            Q(fecha_creacion__icontains=search_query)
        ).distinct()

    # Lista de técnicos para el selector en la plantilla: si es técnico, solo incluirse a sí mismo
    if is_tech_user:
        users = CustomUser.objects.filter(id=request.user.id)
    else:
        users = CustomUser.objects.filter(rol__in=['tecnico', 'administrador'])

    return render(request, 'reports/historial_equipos.html', {'registros': registros, 'users': users, 'is_tech_user': is_tech_user})

@login_required
def get_equipo_details(request, registro_id):
    try:
        registro = FichaEntrada.objects.get(id=registro_id)
        # Si el usuario es técnico, solo puede ver detalles de las fichas que tiene asignadas
        is_tech_user = (getattr(request.user, 'rol', None) == 'tecnico') and (not request.user.is_superuser)
        if is_tech_user and registro.tecnico_asignado != request.user:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        data = {
            'codigo': registro.codigo,
            'tipo_equipo': registro.get_tipo_equipo_display(),
            'marca': registro.marca,
            'modelo': registro.modelo,
            'numero_serie': registro.numero_serie,
            'ubicacion': registro.get_ubicacion_display(),
            'dependencia': registro.dependencia,
            'descripcion': registro.descripcion,
            'nombre_cliente': registro.nombre_cliente,
            'apellido_cliente': registro.apellido_cliente,
            'cedula_cliente': registro.cedula_cliente,
            'departamento_cliente': registro.departamento_cliente,
            'telefono_cliente': registro.telefono_cliente,
            'correo_cliente': registro.correo_cliente,
            'descripcion_falla': registro.descripcion_falla,
            'tipo_falla': registro.get_tipo_falla_display(),
            'observaciones': registro.observaciones,
            'fecha_creacion': registro.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S"),
            'tecnico_asignado': registro.tecnico_asignado.get_full_name() if registro.tecnico_asignado else '',
        }
        return JsonResponse(data)
    except ObjectDoesNotExist:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def asignar_tecnico(request, registro_id):
    if request.method == 'POST':
        tecnico_id = request.POST.get('tecnico')
        tecnico = CustomUser.objects.get(id=tecnico_id)
        registro = FichaEntrada.objects.get(id=registro_id)
        registro.tecnico_asignado = tecnico  # Asignar la instancia del usuario
        registro.save()
        return redirect('reports:historial_equipos')  # Asegúrate de usar el nombre correcto de la URL
    # Obtener los usuarios con roles de tecnico o administrador
    tecnicos = CustomUser.objects.filter(rol__in=['tecnico', 'administrador'])
    return render(request, 'reports/asignar_tecnico.html', {'tecnicos': tecnicos})

def exportar_datos(request):
    search_query = request.GET.get('search', '')
    # Base queryset
    registros = FichaEntrada.objects.all()
    # Si el usuario es técnico (y no es superuser), limitar a sus fichas asignadas
    is_tech_user = (getattr(request.user, 'rol', None) == 'tecnico') and (not request.user.is_superuser)
    if is_tech_user:
        registros = registros.filter(tecnico_asignado=request.user)

    # Aplicar búsqueda sobre el queryset ya filtrado
    if search_query:
        registros = registros.filter(
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(nombre_cliente__icontains=search_query) |
            Q(apellido_cliente__icontains=search_query) |
            Q(departamento_cliente__icontains=search_query) |
            Q(tipo_equipo__icontains=search_query) |
            Q(codigo__icontains=search_query)
        ).distinct()

    # 1. Crear un nuevo libro de trabajo de Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 2. Definir el estilo de fuente en negrita
    bold_font = Font(bold=True)

    # 3. Encabezados (Fila 1)
    headers = [
        'Código',
        'Tipo de Equipo',
        'Marca',
        'Modelo',
        'Cliente',
        'Tipo de Falla',
        'Fecha de Creación',
    ]

    # Escribir los encabezados en la primera fila y aplicar negrita
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_title)
        cell.font = bold_font  # <--- Aplicar el estilo de negrita

    # 4. Escribir los datos (A partir de la Fila 2)
    row_num = 2
    for registro in registros:
        sheet.cell(row=row_num, column=1, value=registro.codigo)
        sheet.cell(row=row_num, column=2, value=registro.get_tipo_equipo_display())
        sheet.cell(row=row_num, column=3, value=registro.marca)
        sheet.cell(row=row_num, column=4, value=registro.modelo)
        sheet.cell(row=row_num, column=5, value=f"{registro.nombre_cliente} {registro.apellido_cliente}")
        sheet.cell(row=row_num, column=6, value=registro.get_tipo_falla_display())
        sheet.cell(row=row_num, column=7, value=registro.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'))
        row_num += 1

    # 5. Crear la respuesta HTTP para el archivo XLSX
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Crear la respuesta HTTP con el tipo de contenido correcto para Excel
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="historial_equipos.xlsx"'

    return response

@login_required
def reporte_estadisticas(request):
    # Construir queryset base aplicando filtros opcionales provistos en GET
    def _apply_stats_filters(req):
        qs = FichaEntrada.objects.all()
        search = (req.GET.get('search') or '').strip()
        if not search:
            return qs

        q = Q()
        # Búsqueda libre sobre campos relevantes
        q |= Q(codigo__icontains=search)
        q |= Q(nombre_cliente__icontains=search)
        q |= Q(apellido_cliente__icontains=search)
        q |= Q(cedula_cliente__icontains=search)
        q |= Q(marca__icontains=search)
        q |= Q(modelo__icontains=search)
        # sede/ubicación (ej: 'barinas')
        q |= Q(ubicacion__icontains=search)
        # número de serie y contactos/cliente
        q |= Q(numero_serie__icontains=search)
        q |= Q(correo_cliente__icontains=search)
        q |= Q(telefono_cliente__icontains=search)
        q |= Q(dependencia__icontains=search)
        q |= Q(departamento_cliente__icontains=search)
        q |= Q(tipo_equipo__icontains=search)
        # si el usuario escribió la etiqueta libre del tipo/falla cuando usó 'otro'
        q |= Q(tipo_equipo_otro__icontains=search)
        q |= Q(tipo_falla__icontains=search)
        q |= Q(tipo_falla_otro__icontains=search)
        q |= Q(descripcion_falla__icontains=search)
        # descripción general/observaciones del equipo
        q |= Q(descripcion__icontains=search)
        q |= Q(observaciones__icontains=search)
        # técnico - buscar por nombre/apellido/username
        q |= Q(tecnico_asignado__first_name__icontains=search)
        q |= Q(tecnico_asignado__last_name__icontains=search)
        q |= Q(tecnico_asignado__username__icontains=search)

        # Intentar interpretar search como fecha (yyyy-mm-dd o dd/mm/yyyy)
        try:
            # primero yyyy-mm-dd
            fecha = datetime.strptime(search, '%Y-%m-%d').date()
            q |= Q(fecha_creacion__date=fecha)
        except Exception:
            try:
                fecha = datetime.strptime(search, '%d/%m/%Y').date()
                q |= Q(fecha_creacion__date=fecha)
            except Exception:
                pass

        return qs.filter(q).distinct()

    base_qs = _apply_stats_filters(request)

    # Si el usuario es técnico (y no es superuser), limitar el queryset a las fichas
    # asignadas a ese técnico para que tanto tablas como gráficos muestren solo su información.
    is_tech_user = (getattr(request.user, 'rol', None) == 'tecnico') and (not request.user.is_superuser)
    if is_tech_user:
        base_qs = base_qs.filter(tecnico_asignado=request.user)

    # Obtenemos estadisticas y metricas relevantes sobre el queryset filtrado
    total_equipos = base_qs.count()
    equipos_por_tipo = base_qs.values('tipo_equipo').annotate(count=Count('tipo_equipo'))
    equipos_por_estado = base_qs.values('estado').annotate(count=Count('estado'))
    # Conteo por técnico (usamos nombre completo cuando esté disponible)
    equipos_por_tecnico_qs = base_qs.values(
        'tecnico_asignado__id',
        'tecnico_asignado__first_name',
        'tecnico_asignado__last_name'
    ).annotate(count=Count('id')).order_by('-count')

    # Preparar datos para los gráficos
    # Reparaciones por mes (año actual, meses 1-12)
    # Permitir seleccionar año/mes desde GET (para ver reparaciones en un mes específico)
    current_year = timezone.now().year
    # leer parámetros opcionales
    month_param = request.GET.get('month')
    year_param = request.GET.get('year')
    try:
        year_int = int(year_param) if year_param else current_year
    except Exception:
        year_int = current_year

    # Usar base_qs para calcular reparaciones por mes del año seleccionado
    reparaciones_mes_qs = (
        base_qs.filter(fecha_creacion__year=year_int)
        .annotate(month=TruncMonth('fecha_creacion'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Si se seleccionó mes específico, calcular el conteo para ese mes
    rep_count_for_month = None
    selected_month = None
    try:
        if month_param:
            month_int = int(month_param)
            if 1 <= month_int <= 12:
                selected_month = month_int
                rep_count_for_month = base_qs.filter(fecha_creacion__year=year_int, fecha_creacion__month=month_int).count()
    except Exception:
        rep_count_for_month = None

    # Mapeo de mes (1-12) a conteo
    meses_map = {item['month'].month: item['count'] for item in reparaciones_mes_qs}
    meses_es = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    reparaciones_por_mes_labels = meses_es
    reparaciones_por_mes_data = [meses_map.get(i, 0) for i in range(1,13)]

    # Preparar datos por tecnico
    reparaciones_por_tecnico_labels = []
    reparaciones_por_tecnico_data = []
    for item in equipos_por_tecnico_qs:
        first = item.get('tecnico_asignado__first_name') or ''
        last = item.get('tecnico_asignado__last_name') or ''
        if first or last:
            label = f"{first} {last}".strip()
        else:
            # Si no hay técnico asignado o datos, usar 'Sin asignar' o username nulo
            label = 'Sin asignar'
        reparaciones_por_tecnico_labels.append(label)
        reparaciones_por_tecnico_data.append(item.get('count', 0))
    
    # tiempo estimado promedio de reparacion, no sabemos cuanto es el tiempo pero supongamos que son 2 semanas
    tiempo_estimado_reparacion = 14 # dias
    context = {
        'total_equipos': total_equipos,
        'equipos_por_tipo': list(equipos_por_tipo),
        'equipos_por_estado': list(equipos_por_estado),
        'equipos_por_tecnico': list(equipos_por_tecnico_qs),
        'tiempo_promedio_reparacion': tiempo_estimado_reparacion,
        # Datos para gráficos (serializados en JSON en la plantilla)
        'reparaciones_por_mes': {
            'labels': reparaciones_por_mes_labels,
            'data': reparaciones_por_mes_data,
        },
        'reparaciones_por_tecnico': {
            'labels': reparaciones_por_tecnico_labels,
            'data': reparaciones_por_tecnico_data,
        },
        # selección de mes/año
        'selected_month': selected_month,
        'selected_year': year_int,
        'rep_count_for_month': rep_count_for_month,
        'available_years': list(range(current_year, current_year-6, -1)),
        'is_tech_user': is_tech_user,
        # JSON strings para inyectar de forma segura en JS
        'reparaciones_por_mes_json': json.dumps({
            'labels': reparaciones_por_mes_labels,
            'data': reparaciones_por_mes_data,
        }),
        'reparaciones_por_tecnico_json': json.dumps({
            'labels': reparaciones_por_tecnico_labels,
            'data': reparaciones_por_tecnico_data,
        }),
    }
    return render(request, 'statistics/admin_statistics.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def add_seguimiento(request, registro_id):
    # obtenemos la ficha de entrada correspondiente
    ficha = get_object_or_404(FichaEntrada, id=registro_id) 
    
    # Si el metodo es Get, mostramos el formulario
    if request.method == 'GET':
        # Si el usuario es técnico, pre-asignar el técnico en el formulario y bloquear la selección en el template
        is_tech_user = (getattr(request.user, 'rol', None) == 'tecnico') and (not request.user.is_superuser)
        if is_tech_user:
            form = SeguimientoForm(initial={'ficha': ficha, 'tecnico': request.user})
        else:
            form = SeguimientoForm(initial={'ficha': ficha})
        html = render_to_string('reports/seguimiento_form_partial.html', {'form': form, 'ficha': ficha}, request=request)
        return JsonResponse({'html': html})
    
    # Si el metodo no es get y es post, procesamos el formulario (incluye archivos)
    form = SeguimientoForm(request.POST, request.FILES)
    logger.info('add_seguimiento: usuario=%s registro=%s archivos=%s', request.user.username, registro_id, list(request.FILES.keys()))

    if not form.is_valid():
        logger.warning('Formulario Seguimiento no válido: %s', form.errors.as_json())
        html = render_to_string('reports/seguimiento_form_partial.html', {'form': form, 'ficha': ficha}, request=request)
        return JsonResponse({'success': False, 'html': html, 'errors': form.errors.get_json_data()}, status=400)

    # Leemos los datos validados del formulario (no guardamos un nuevo objeto aquí)
    cleaned = form.cleaned_data
    estado = cleaned.get('estado')
    # Forzar progreso según el estado (backend authoritative)
    estado_to_progress = {
        'recepcion': 5,
        'diagnostico': 20,
        'reparacion': 60,
        'pruebas': 85,
        'listo': 95,
        'entregado': 100,
        'otro': 0,
    }
    progreso = estado_to_progress.get(estado, 0)
    descripcion = cleaned.get('descripcion') or ''
    # Si el usuario es técnico, forzamos que el técnico del seguimiento sea el mismo usuario
    if (getattr(request.user, 'rol', None) == 'tecnico') and (not request.user.is_superuser):
        tecnico_instance = request.user
    else:
        tecnico_instance = cleaned.get('tecnico') if 'tecnico' in cleaned else None

    video_file = request.FILES.get('video')
    video_url = None
    video_trimmed = False
    video_trim_error = None

    try:
        with transaction.atomic():
            # Bloquear y obtener el seguimiento existente si existe
            seguimiento = Seguimiento.objects.select_for_update().filter(ficha=ficha).first()
            # Validación: no permitir registrar un estado que ya exista en el timeline
            if seguimiento:
                try:
                    existing_timeline = list(seguimiento.timeline or [])
                    used_estados = set([(ev.get('estado') or '').strip().lower() for ev in existing_timeline if ev.get('estado')])
                    new_estado = (estado or '').strip().lower()
                    if new_estado and new_estado in used_estados:
                        # Determinar siguiente estado sugerido según la secuencia lógica
                        sequence = ['recepcion', 'diagnostico', 'reparacion', 'pruebas', 'listo', 'entregado']
                        choice_map = dict(Seguimiento.ESTADO_CHOICES)
                        next_state = None
                        for s in sequence:
                            if s not in used_estados:
                                next_state = s
                                break

                        # Preparar mensaje legible
                        used_display = choice_map.get(new_estado, new_estado)
                        if next_state:
                            next_display = choice_map.get(next_state, next_state)
                            msg = f"No puede usar el estado '{used_display}' porque ya fue usado en el timeline. Seleccione el siguiente estado: '{next_display}'."
                        else:
                            msg = f"No puede usar el estado '{used_display}' porque ya fue usado en el timeline. No hay estados siguientes disponibles."

                        form.add_error(None, msg)
                        html = render_to_string('reports/seguimiento_form_partial.html', {'form': form, 'ficha': ficha}, request=request)
                        return JsonResponse({'success': False, 'html': html, 'message': msg}, status=400)
                except Exception:
                    logger.exception('Error verificando estados existentes en el timeline para ficha %s', ficha.id)
            if not seguimiento:
                # Crear un seguimiento base
                seguimiento = Seguimiento.objects.create(
                    ficha=ficha,
                    estado=estado or 'recepcion',
                    progreso=progreso,
                    descripcion=descripcion,
                    tecnico=tecnico_instance if tecnico_instance else None,
                )
            else:
                # Actualizar campos principales del seguimiento existente
                if estado:
                    seguimiento.estado = estado
                seguimiento.progreso = progreso
                if descripcion:
                    seguimiento.descripcion = descripcion
                if tecnico_instance:
                    seguimiento.tecnico = tecnico_instance
                seguimiento.save()

            # Si se subió un video, guardarlo en storage con nombre único
            if video_file:
                ext = (video_file.name.split('.')[-1] or 'mp4').lower()
                filename = f"seguimientos/videos/{uuid4().hex}.{ext}"
                saved_path = default_storage.save(filename, video_file)
                try:
                    video_url = default_storage.url(saved_path)
                except Exception:
                    video_url = os.path.join(settings.MEDIA_URL, saved_path)

                # Intentar recortar si storage provee un path local
                try:
                    local_path = default_storage.path(saved_path)
                    from moviepy.editor import VideoFileClip
                    clip = VideoFileClip(local_path)
                    end_t = min(10, clip.duration)
                    if clip.duration > end_t:
                        tmp_path = local_path + '.tmp.mp4'
                        clip.subclip(0, end_t).write_videofile(tmp_path, codec='libx264', audio_codec='aac', logger=None)
                        clip.close()
                        os.replace(tmp_path, local_path)
                    else:
                        clip.close()
                    video_trimmed = True
                except NotImplementedError:
                    # Storage no provee path local (p.ej. S3) — no intentamos recortar aquí
                    video_trim_error = 'storage_has_no_local_path'
                except Exception as e:
                    logger.exception('Error recortando video: %s', e)
                    video_trim_error = str(e)

            # Construir evento y añadir al timeline
            evento = {
                'fecha': timezone.now().isoformat(),
                'titulo': dict(Seguimiento.ESTADO_CHOICES).get(estado, estado),
                'descripcion': descripcion,
                'estado': estado,
                'progreso': progreso,
                'tecnico': tecnico_instance.get_full_name() if tecnico_instance else '',
                'autor': request.user.get_full_name() or request.user.username,
                'video': video_url,
            }

            existing_timeline = list(seguimiento.timeline or [])
            existing_timeline.append(evento)
            seguimiento.timeline = existing_timeline
            seguimiento.save()

            # Sincronizar el estado del seguimiento con la FichaEntrada relacionada
            # Mapeo entre los estados del modelo Seguimiento y los valores válidos en FichaEntrada
            estado_map = {
                'recepcion': 'recibido',
                'diagnostico': 'diagnostico',
                'reparacion': 'reparacion',
                'pruebas': 'pruebas_calidad',
                'listo': 'listo_entrega',
                'entregado': 'entregado',
            }
            try:
                ficha_estado = estado_map.get((estado or '').strip().lower())
                if ficha_estado:
                    ficha.estado = ficha_estado
                    ficha.save()
                    logger.info('Ficha %s actualizada a estado %s por seguimiento %s', ficha.id, ficha_estado, seguimiento.id)
            except Exception:
                logger.exception('Error actualizando estado de FichaEntrada para ficha %s', getattr(ficha, 'id', None))

            # Enviar correo automático según el estado del evento
            try:
                def _send_event_email(ficha_obj, seguimiento_obj, evento_obj):
                    if not ficha_obj.correo_cliente:
                        logger.info('No hay correo de cliente para ficha %s, no se enviará email', ficha_obj.id)
                        return
                    # Soportar sinónimos de estados para evitar que diferencias en las claves impidan el envio
                    template_map = {
                        'recepcion': 'emails/recepcion_equipo.html',
                        'recibido': 'emails/recepcion_equipo.html',
                        'listo': 'emails/retirar_equipo.html',
                        'listo_entrega': 'emails/retirar_equipo.html',
                        'entregado': 'emails/entregado_equipo.html',
                    }
                    # Normalizar y soportar variantes/typos
                    estado_key = (evento_obj.get('estado') or '').strip().lower()
                    # Corrección rápida de typos comunes
                    if estado_key == 'rececion':
                        estado_key = 'recepcion'
                    tpl = template_map.get(estado_key)
                    if not tpl:
                        logger.info('Estado "%s" no mapeado para envío de email (ficha=%s). Keys disponibles: %s', estado_key, ficha_obj.id, ','.join(template_map.keys()))
                        return
                    subject = f"Estado de su equipo {ficha_obj.codigo}: {evento_obj.get('titulo', '')}"
                    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@example.com')
                    to_email = [ficha_obj.correo_cliente]
                    
                    equipo_ctx = {
                        'codigo': getattr(ficha_obj, 'codigo', ''),
                        # usar display() para mostrar la etiqueta legible del choice
                        'tipo_equipo': ficha_obj.get_tipo_equipo_display() if hasattr(ficha_obj, 'get_tipo_equipo_display') else getattr(ficha_obj, 'tipo_equipo', ''),
                        'marca': getattr(ficha_obj, 'marca', ''),
                        'modelo': getattr(ficha_obj, 'modelo', ''),
                        # campo de estado: preferir el estado del seguimiento si existe, sino el de la ficha
                        'estado_actual': (seguimiento_obj.estado if seguimiento_obj and getattr(seguimiento_obj, 'estado', None) else getattr(ficha_obj, 'estado', 'En espera de diagnóstico')),
                    }
                    # Fecha de entrega: si el seguimiento tiene fecha_estimada/fecha_entrega, usarla; si no, dejar vacío o usar now()
                    fecha_entrega = None
                    if seguimiento_obj and getattr(seguimiento_obj, 'fecha_entrega', None):
                        fecha_entrega = seguimiento_obj.fecha_entrega
                    elif getattr(ficha_obj, 'fecha_entrega', None):
                        fecha_entrega = ficha_obj.fecha_entrega
                    else:
                        fecha_entrega = None
                        
                    context = {
                        'ficha': ficha_obj,
                        'seguimiento': seguimiento_obj,
                        'evento': evento_obj,
                        'equipo': equipo_ctx,
                        'fecha_entrega': fecha_entrega,
                    }
                    # Pasar el `request` para que los context processors (p.ej. `user`) estén disponibles
                    try:
                        html_content = render_to_string(tpl, context, request=request)
                    except Exception:
                        # Fallback: intentar render sin request si algo falla
                        logger.exception('Falló render_to_string con request, intentando sin request')
                        html_content = render_to_string(tpl, context)
                    text_content = strip_tags(html_content)
                    try:
                        logger.info('Intentando enviar correo a %s (ficha=%s, estado=%s, plantilla=%s)', ficha_obj.correo_cliente, ficha_obj.id, estado_key, tpl)
                        # Usar send_mail con html_message; send_mail devuelve el número de mensajes enviados
                        sent = send_mail(
                            subject,
                            text_content,
                            from_email,
                            to_email,
                            fail_silently=False,
                            html_message=html_content,
                        )
                        logger.info('send_mail returned: %s', sent)
                        logger.info('Correo enviado a %s para ficha %s estado=%s', ficha_obj.correo_cliente, ficha_obj.id, estado_key)
                    except Exception as send_ex:
                        # Registrar y encolar el correo para reintento en background
                        logger.exception('Error enviando correo a %s: %s', ficha_obj.correo_cliente, send_ex)
                        try:
                            QueuedEmail.objects.create(
                                to=[ficha_obj.correo_cliente],
                                subject=subject,
                                body_text=text_content,
                                body_html=html_content,
                                attempts=0,
                                last_error=str(send_ex),
                            )
                            logger.info('Correo encolado para reintento a %s (ficha=%s)', ficha_obj.correo_cliente, ficha_obj.id)
                        except Exception:
                            logger.exception('No se pudo encolar el correo para ficha %s', ficha_obj.id)
                    

                # Llamada al helper (no bloquear si falla)
                _send_event_email(ficha, seguimiento, evento)
            except Exception:
                logger.exception('Error en la rutina de envío de correo para el evento')

    except Exception as e:
        logger.exception('Error procesando add_seguimiento: %s', e)
        return JsonResponse({'success': False, 'message': 'Error interno procesando el seguimiento.'}, status=500)

    return JsonResponse({'success': True, 'message': 'Evento añadido al timeline', 'event': evento, 'seguimiento_id': seguimiento.id, 'video_trimmed': video_trimmed, 'video_trim_error': video_trim_error})


def timelines_by_cedula(request):
    """Devuelve JSON con las fichas del cliente identificado por `cedula` y sus timelines.

    Parámetros GET:
    - cedula: número de cédula del cliente

    Respuesta JSON:
    { success: true, cedula: '...', fichas: [ { id, codigo, tipo_equipo, modelo, estado, progreso, fecha_ingreso, fecha_estimada, timeline: [...] }, ... ] }
    """
    cedula = request.GET.get('cedula', '').strip()
    if not cedula:
        return JsonResponse({'success': False, 'message': 'Parámetro cedula requerido.'}, status=400)

    try:
        fichas = FichaEntrada.objects.filter(cedula_cliente=cedula)
        result = []
        for ficha in fichas:
            # Obtener el seguimiento asociado (si existe)
            seguimiento = Seguimiento.objects.filter(ficha=ficha).order_by('-fecha_ingreso').first()
            timeline = []
            if seguimiento:
                # Asegurarse de que timeline sea lista
                timeline = list(seguimiento.timeline or [])
                # Normalizar campos: si un evento tiene ruta relativa de video, dejarla como viene
            # Formatear datos de ficha para presentar en el modal
            item = {
                'id': ficha.id,
                'codigo': ficha.codigo,
                'tipo_equipo': ficha.get_tipo_equipo_display(),
                'modelo': ficha.modelo,
                'estado': seguimiento.estado if seguimiento else ficha.estado,
                'progreso': seguimiento.progreso if seguimiento else 0,
                'fechaIngreso': ficha.fecha_creacion.strftime('%d %b %Y'),
                'fechaEstimada': seguimiento.fecha_estimada.strftime('%d %b %Y') if (seguimiento and seguimiento.fecha_estimada) else '',
                'timeline': timeline,
            }
            result.append(item)

        return JsonResponse({'success': True, 'cedula': cedula, 'fichas': result})
    except Exception as e:
        logger.exception('Error timelines_by_cedula: %s', e)
        return JsonResponse({'success': False, 'message': 'Error interno al obtener timelines.'}, status=500)
        
