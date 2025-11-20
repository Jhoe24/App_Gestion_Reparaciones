# =====================================================
# ARCHIVO: apps/users/views.py
# =====================================================

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q

from .models import CustomUser, LogAcceso
from apps.authentication.mixins import AdminRequiredMixin
from apps.authentication.forms import CustomUserCreationForm, CustomUserChangeForm
from django.http import HttpResponse, JsonResponse
import csv
import io
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors
except Exception:
    SimpleDocTemplate = None
from django.utils import timezone
from django.db.models import Count
from datetime import timedelta


class UserListView(AdminRequiredMixin, ListView):
    """Vista para listar usuarios (solo admin)"""
    model = CustomUser
    template_name = 'users/list_user_admin.html'
    context_object_name = 'users'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = CustomUser.objects.all().order_by('last_name', 'first_name')
        
        # Filtro por búsqueda
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(cedula__icontains=search) |
                Q(username__icontains=search)
            )
        
        # Filtro por rol
        rol = self.request.GET.get('rol')
        if rol:
            queryset = queryset.filter(rol=rol)
        
        # Filtro por estado
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(activo=True)
        elif status == 'inactive':
            queryset = queryset.filter(activo=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['rol_selected'] = self.request.GET.get('rol', '')
        context['status_selected'] = self.request.GET.get('status', '')
        context['roles'] = CustomUser.ROLES
        return context


class UserDetailView(AdminRequiredMixin, DetailView):
    """Vista de detalle de usuario"""
    model = CustomUser
    template_name = 'users/user_detail.html'
    context_object_name = 'user_obj'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recent_logs'] = LogAcceso.objects.filter(
            usuario=self.object
        )[:10]
        return context


class UserCreateView(AdminRequiredMixin, CreateView):
    """Vista para crear usuario"""
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('users:user_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Usuario creado exitosamente.')
        return super().form_valid(form)


class UserUpdateView(AdminRequiredMixin, UpdateView):
    """Vista para editar usuario"""
    model = CustomUser
    form_class = CustomUserChangeForm  # Usar el nuevo formulario que incluye todos los campos
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('users:user_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Usuario actualizado exitosamente.')
        return super().form_valid(form)


class UserDeleteView(AdminRequiredMixin, DeleteView):
    """Vista para eliminar usuario"""
    model = CustomUser
    template_name = 'users/user_confirm_delete.html'
    success_url = reverse_lazy('users:user_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Variable para cambiar el texto en la plantilla
        context['is_deactivation'] = True
        return context

    def delete(self, request, *args, **kwargs):
        """En lugar de borrar, desactiva al usuario (borrado lógico)."""
        self.object = self.get_object()
        success_url = self.get_success_url()
        if self.object == request.user:
            messages.error(request, 'No puedes desactivar tu propia cuenta.')
            return redirect(success_url)
 
        self.object.activo = False
        self.object.save()
        messages.success(request, f'El usuario "{self.object.get_full_name()}" ha sido desactivado exitosamente.')
        return redirect(success_url)


@login_required
def toggle_user_status(request, pk):
    """Función para activar/desactivar usuario"""
    if not request.user.is_administrador:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('equipment:user_dashboard')
    
    user = get_object_or_404(CustomUser, pk=pk)
    
    if user == request.user:
        messages.error(request, 'No puedes desactivar tu propia cuenta.')
        return redirect('users:user_list')
    
    user.activo = not user.activo
    user.save()
    
    status = "activado" if user.activo else "desactivado"
    messages.success(request, f'Usuario {user.get_full_name()} {status} exitosamente.')
    
    return redirect('users:user_list')

@login_required
def profile_view(request):
    """Vista para que el usuario vea y edite su propio perfil"""
    user = request.user
    
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            return redirect('users:profile')
    else:
        form = CustomUserChangeForm(instance=user)
    
    return render(request, 'profile/profile_admin.html', {'form': form, 'user_obj': user})

@login_required
def list_user_view(request):
    """Vista para que el admin vea la lista de usuarios"""
    users = CustomUser.objects.all().order_by('fecha_creacion')
    
    context = {
        'users': users
    }
    return render(request, 'users/list_user_admin.html', context)

@login_required
def new_user_view(request):
    """Vista para que el admin cree un nuevo usuario"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('users:list_user_admin')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'users/new_user.html', {'form': form})


class LogAccesoListView(AdminRequiredMixin, ListView):
    """Lista los logs de acceso (solo administradores)."""
    model = LogAcceso
    template_name = 'users/logs.html'
    context_object_name = 'logs'
    paginate_by = 50

    def get_queryset(self):
        queryset = LogAcceso.objects.select_related('usuario').all().order_by('-fecha_hora')

        # filtros simples: usuario, accion, rango de fechas
        usuario = self.request.GET.get('usuario')
        accion = self.request.GET.get('accion')
        desde = self.request.GET.get('desde')
        hasta = self.request.GET.get('hasta')

        if usuario:
            queryset = queryset.filter(usuario__id=usuario)

        if accion:
            queryset = queryset.filter(accion=accion)

        if desde:
            try:
                queryset = queryset.filter(fecha_hora__date__gte=desde)
            except Exception:
                pass

        if hasta:
            try:
                queryset = queryset.filter(fecha_hora__date__lte=hasta)
            except Exception:
                pass

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['acciones'] = LogAcceso.ACCIONES
        context['usuarios'] = CustomUser.objects.order_by('last_name', 'first_name')
        context['filtros'] = {
            'usuario': self.request.GET.get('usuario', ''),
            'accion': self.request.GET.get('accion', ''),
            'desde': self.request.GET.get('desde', ''),
            'hasta': self.request.GET.get('hasta', ''),
        }
        return context


@login_required
def logs_json(request):
    """Devuelve datos agregados para gráficos: conteo de acciones en los últimos N días."""
    if not request.user.is_administrador:
        return JsonResponse({'error': 'Permiso denegado'}, status=403)

    days = int(request.GET.get('days', 30))
    since = timezone.now() - timedelta(days=days)

    data = (
        LogAcceso.objects.filter(fecha_hora__gte=since)
        .values('accion')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Formato: [{accion: 'login', count: 10}, ...]
    result = [{'accion': item['accion'], 'count': item['count']} for item in data]
    return JsonResponse({'data': result})


@login_required
def export_logs_csv(request):
    """Exporta los logs filtrados a CSV."""
    if not request.user.is_administrador:
        return HttpResponse('Permiso denegado', status=403)

    # Reutilizamos los mismos parámetros de filtrado que la lista
    queryset = LogAcceso.objects.select_related('usuario').all().order_by('-fecha_hora')
    usuario = request.GET.get('usuario')
    accion = request.GET.get('accion')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    if usuario:
        queryset = queryset.filter(usuario__id=usuario)
    if accion:
        queryset = queryset.filter(accion=accion)
    if desde:
        try:
            queryset = queryset.filter(fecha_hora__date__gte=desde)
        except Exception:
            pass
    if hasta:
        try:
            queryset = queryset.filter(fecha_hora__date__lte=hasta)
        except Exception:
            pass

    # Construir respuesta CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="logs_acceso.csv"'

    writer = csv.writer(response)
    writer.writerow(['Usuario', 'Accion', 'Fecha Hora', 'Direccion IP', 'User Agent'])

    for log in queryset:
        usuario_nombre = log.usuario.get_full_name() if log.usuario else 'Desconocido'
        writer.writerow([usuario_nombre, log.get_accion_display(), log.fecha_hora.isoformat(), log.direccion_ip or '', (log.user_agent or '').replace('\n', ' ')])

    return response


@login_required
def export_logs_excel(request):
    """Exporta los logs filtrados a Excel (.xlsx)."""
    if not request.user.is_administrador:
        return HttpResponse('Permiso denegado', status=403)

    if Workbook is None:
        return HttpResponse('Dependencia missing: openpyxl no instalada', status=500)

    # Reutilizamos filtros
    queryset = LogAcceso.objects.select_related('usuario').all().order_by('-fecha_hora')
    usuario = request.GET.get('usuario')
    accion = request.GET.get('accion')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    if usuario:
        queryset = queryset.filter(usuario__id=usuario)
    if accion:
        queryset = queryset.filter(accion=accion)
    if desde:
        try:
            queryset = queryset.filter(fecha_hora__date__gte=desde)
        except Exception:
            pass
    if hasta:
        try:
            queryset = queryset.filter(fecha_hora__date__lte=hasta)
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = 'Logs'

    headers = ['Usuario', 'Accion', 'Fecha Hora', 'Direccion IP', 'User Agent']
    ws.append(headers)

    for log in queryset:
        usuario_nombre = log.usuario.get_full_name() if log.usuario else 'Desconocido'
        ws.append([usuario_nombre, log.get_accion_display(), log.fecha_hora.strftime('%Y-%m-%d %H:%M:%S'), log.direccion_ip or '', (log.user_agent or '').replace('\n', ' ')])

    # Ajustar ancho de columnas
    for i, column_cells in enumerate(ws.columns, 1):
        length = max((len(str(cell.value)) for cell in column_cells), default=0)
        ws.column_dimensions[get_column_letter(i)].width = min(max(length + 2, 10), 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="logs_acceso.xlsx"'
    return response


@login_required
def export_logs_pdf(request):
    """Exporta los logs filtrados a PDF (tabla simple)."""
    if not request.user.is_administrador:
        return HttpResponse('Permiso denegado', status=403)

    if SimpleDocTemplate is None:
        return HttpResponse('Dependencia missing: reportlab no instalada', status=500)

    # Reutilizamos filtros
    queryset = LogAcceso.objects.select_related('usuario').all().order_by('-fecha_hora')
    usuario = request.GET.get('usuario')
    accion = request.GET.get('accion')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    if usuario:
        queryset = queryset.filter(usuario__id=usuario)
    if accion:
        queryset = queryset.filter(accion=accion)
    if desde:
        try:
            queryset = queryset.filter(fecha_hora__date__gte=desde)
        except Exception:
            pass
    if hasta:
        try:
            queryset = queryset.filter(fecha_hora__date__lte=hasta)
        except Exception:
            pass

    # Construir datos para la tabla
    headers = ['Usuario', 'Accion', 'Fecha Hora', 'Direccion IP']
    data = [headers]
    for log in queryset:
        usuario_nombre = log.usuario.get_full_name() if log.usuario else 'Desconocido'
        data.append([usuario_nombre, log.get_accion_display(), log.fecha_hora.strftime('%Y-%m-%d %H:%M:%S'), log.direccion_ip or ''])
    # agremamos un titulo
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20, title="Registros de Acceso")
    table = Table(data, repeatRows=1)
    
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E6DA4')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
    ])
    table.setStyle(style)
    elems = [table]
    doc.build(elems)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="logs_acceso.pdf"'
    return response